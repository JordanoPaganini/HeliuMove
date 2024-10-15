from django.shortcuts import render, get_object_or_404
from .models import *
from django.core.paginator import Paginator
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse


def sistemas(request, slug):
    sistema = get_object_or_404(Sistemas, slug=slug)
    dados_list = Dados.objects.filter(sistema=sistema).order_by('-datatime')
    
    # Paginação
    paginator = Paginator(dados_list, 25)  # 25 dados por página
    page_number = request.GET.get('page')  # Obtém o número da página atual
    dados = paginator.get_page(page_number)  # Obtém os dados para a página atual

    data = {
        'nome': sistema.nome,
        'dados': dados,
        'paginator': paginator,
    }
    return render(request, 'sistemas.html', data)

def export(request, slug):
    # Buscar o sistema pelo slug
    sistema = Sistemas.objects.get(slug=slug)
    
    # Obter todos os dados relacionados a esse sistema
    dados = Dados.objects.filter(sistema=sistema).order_by('-datatime')
    
    # Criar um arquivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = sistema.nome
    
    # Estilos para formatação
    bold_font = Font(bold=True)  # Negrito
    center_alignment = Alignment(horizontal="center", vertical="center")  # Alinhamento centralizado
    
    # Preencher as cores para as linhas
    fill_gray_header = PatternFill(start_color="bdbdbd", end_color="bdbdbd", fill_type="solid")  # Cor do cabeçalho
    fill_gray_alt = PatternFill(start_color="f3f3f3", end_color="f3f3f3", fill_type="solid")  # Cor cinza claro
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Cor branca

    
    # Adicionar cabeçalhos ao Excel com formatação em negrito, centralizado e com borda
    headers = ['Data/Hora', 'Temperatura (°C)', 'Tensão (V)', 'Corrente (A)']
    ws.append(headers)
    
    # Aplicar formatação aos cabeçalhos (negrito, alinhado, borda e cor de fundo)
    for cell in ws[1]:
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.fill = fill_gray_header  # Cor de fundo
        cell.border = None

    # Preencher o arquivo com os dados do sistema
    for index, dado in enumerate(dados, start=2):  # Inicia da linha 2 (depois do cabeçalho)
        # Alternar cor das linhas
        if index % 2 == 0:
            fill = fill_gray_alt  # Linhas pares com fundo cinza claro
        else:
            fill = fill_white  # Linhas ímpares com fundo branco
        
        # Adicionar os dados nas células
        ws.append([
            dado.datatime.strftime('%d/%m/%y %H:%M:%S'),  # Data formatada como DD/MM/YY HH:MM:SS
            f"{dado.temp} °C",  # Temperatura com unidade
            f"{dado.voltage} V",  # Tensão com unidade
            f"{dado.current} A"   # Corrente com unidade
        ])
        
        # Aplicar formatação (alinhamento, borda e cor de fundo) nas células da linha atual
        for cell in ws[index]:
            cell.alignment = center_alignment  # Alinhamento centralizado
            cell.fill = fill  # Cor de fundo alternada
            cell.border = None

    # Ajustar automaticamente a largura das colunas
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Configurar o arquivo para download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={sistema.nome}_dados.xlsx'
    
    # Salvar o arquivo no response
    wb.save(response)
    
    return response